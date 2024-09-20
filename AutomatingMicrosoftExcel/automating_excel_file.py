import openpyxl as xl

def StudentScores(filename):
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 2)
            cell1 = sheet.cell(row, 3)

            mid_sem_30 = cell.value * 0.30 
            end_sem_70 = cell1.value * 0.70

            mid_sem_cell = sheet.cell(row, 4) 
            end_sem_cell = sheet.cell(row, 5) 

            mid_sem_cell.value = mid_sem_30
            end_sem_cell.value = end_sem_70

            initial_total_score = mid_sem_cell.value + end_sem_cell.value
            initial_total_score_cell = sheet.cell(row, 6)
            initial_total_score_cell.value = initial_total_score


        wb.save(filename)



StudentScores('StudentScores.xlsx')
